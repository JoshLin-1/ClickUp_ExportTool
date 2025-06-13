import pandas as pd
from typing import Dict, List, Any
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from config import (
    EXCEL_MAX_SHEET_NAME_LENGTH, EXCEL_COLUMN_MIN_WIDTH, 
    EXCEL_COLUMN_MAX_WIDTH, WORKSPACE_COLOR, LIST_COLOR, 
    TASK_COLOR, TOTAL_COLOR, EXCEL_HEADERS
)
from utils.string_utils import make_safe_sheet_name, calculate_display_width


class ExcelExporter:
    """Handles Excel export functionality"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.writer = None
    
    def export(
        self,
        hierarchical_summary: List[Dict[str, Any]],
        user_summary: List[Dict[str, Any]],
        all_user_entries: Dict[str, Dict[str, Any]],
        progress_callback=None
    ):
        """Export all data to Excel file"""
        with pd.ExcelWriter(self.file_path, engine='openpyxl') as self.writer:
            # Export hierarchical summary
            self._export_hierarchical_summary(hierarchical_summary)
            if progress_callback:
                progress_callback(1)
            
            # Export user summary
            self._export_user_summary(user_summary)
            if progress_callback:
                progress_callback(2)
            
            # Export individual user tabs
            self._export_user_tabs(all_user_entries, progress_callback)
    
    def _export_hierarchical_summary(self, data: List[Dict[str, Any]]):
        """Export hierarchical summary to Excel"""
        df = pd.DataFrame(data)
        df.to_excel(self.writer, sheet_name='Hierarchical Summary', index=False)
        
        self._auto_adjust_columns('Hierarchical Summary', df)
        self._apply_hierarchical_formatting('Hierarchical Summary', df)
    
    def _export_user_summary(self, data: List[Dict[str, Any]]):
        """Export user summary to Excel"""
        df = pd.DataFrame(data)
        df.to_excel(self.writer, sheet_name='User Summary', index=False)
        self._auto_adjust_columns('User Summary', df)
    
    def _export_user_tabs(
        self,
        all_user_entries: Dict[str, Dict[str, Any]],
        progress_callback=None
    ):
        """Export individual user tabs"""
        for i, (username, user_data) in enumerate(all_user_entries.items()):
            entries = user_data['entries']
            safe_name = make_safe_sheet_name(username)
            
            if not entries:
                # Create empty sheet
                df = pd.DataFrame(columns=list(EXCEL_HEADERS.values()))
            else:
                # Sort entries by date
                sorted_entries = sorted(entries, key=lambda x: x.start)
                
                # Convert to Excel rows
                rows = [entry.to_excel_row(username) for entry in sorted_entries]
                df = pd.DataFrame(rows)
            
            df.to_excel(self.writer, sheet_name=safe_name, index=False)
            self._auto_adjust_columns(safe_name, df)
            self._make_urls_clickable(safe_name, df)
            
            if progress_callback:
                progress_callback(i + 3)
    
    def _auto_adjust_columns(self, sheet_name: str, dataframe: pd.DataFrame):
        """Auto-adjust column widths"""
        try:
            worksheet = self.writer.sheets[sheet_name]
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        length = calculate_display_width(str(cell.value))
                        max_length = max(max_length, length)
                
                adjusted_width = min(
                    max(max_length + 2, EXCEL_COLUMN_MIN_WIDTH),
                    EXCEL_COLUMN_MAX_WIDTH
                )
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            print(f"Warning: Could not auto-adjust columns for {sheet_name}: {e}")
    
    def _apply_hierarchical_formatting(self, sheet_name: str, dataframe: pd.DataFrame):
        """Apply formatting to hierarchical summary"""
        try:
            worksheet = self.writer.sheets[sheet_name]
            
            # Define styles
            styles = {
                'WORKSPACE': {
                    'fill': PatternFill(start_color=WORKSPACE_COLOR, end_color=WORKSPACE_COLOR, fill_type="solid"),
                    'font': Font(bold=True, color="FFFFFF")
                },
                'LIST': {
                    'fill': PatternFill(start_color=LIST_COLOR, end_color=LIST_COLOR, fill_type="solid"),
                    'font': Font(bold=True, color="FFFFFF")
                },
                'TASK': {
                    'fill': PatternFill(start_color=TASK_COLOR, end_color=TASK_COLOR, fill_type="solid"),
                    'font': Font(bold=True)
                },
                'GRAND TOTAL': {
                    'fill': PatternFill(start_color=TOTAL_COLOR, end_color=TOTAL_COLOR, fill_type="solid"),
                    'font': Font(bold=True, color="FFFFFF")
                }
            }
            
            # Apply formatting
            for row_num in range(2, len(dataframe) + 2):
                level = worksheet[f'A{row_num}'].value
                
                if level in styles:
                    style = styles[level]
                    for col_num in range(1, len(dataframe.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.fill = style['fill']
                        cell.font = style['font']
                        
                        # Apply alignment
                        if col_num <= 5:
                            cell.alignment = Alignment(horizontal='left')
                        else:
                            cell.alignment = Alignment(horizontal='right')
                            
        except Exception as e:
            print(f"Warning: Could not apply formatting to {sheet_name}: {e}")
    
    def _make_urls_clickable(self, sheet_name: str, dataframe: pd.DataFrame):
        """Make URLs clickable in Excel"""
        try:
            worksheet = self.writer.sheets[sheet_name]
            
            # Find URL column
            url_col_index = None
            for idx, col in enumerate(dataframe.columns):
                if col == EXCEL_HEADERS['url']:
                    url_col_index = idx + 1
                    break
            
            if url_col_index is None:
                return
            
            url_col_letter = get_column_letter(url_col_index)
            
            # Make URLs clickable
            for row_num in range(2, len(dataframe) + 2):
                cell = worksheet[f'{url_col_letter}{row_num}']
                if cell.value and str(cell.value).startswith('http'):
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0000FF", underline="single")
                    cell.value = "Link"
                    
        except Exception as e:
            print(f"Warning: Could not make URLs clickable for {sheet_name}: {e}")