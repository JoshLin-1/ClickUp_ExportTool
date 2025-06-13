# utils/excel_utils.py
"""Excel formatting and manipulation utilities"""

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, Any, List


def create_header_style() -> Dict[str, Any]:
    """Create style for header rows"""
    return {
        'font': Font(bold=True, size=12),
        'fill': PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        'alignment': Alignment(horizontal='center', vertical='center')
    }


def create_hyperlink_style() -> Font:
    """Create style for hyperlinks"""
    return Font(color="0000FF", underline="single")


def apply_border_to_range(worksheet: Worksheet, start_row: int, end_row: int, 
                         start_col: int, end_col: int):
    """Apply borders to a cell range"""
    from openpyxl.styles import Border, Side
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border


def freeze_panes(worksheet: Worksheet, row: int = 2, column: int = 1):
    """Freeze panes at specified row and column"""
    from openpyxl.utils import get_column_letter
    cell_address = f"{get_column_letter(column)}{row}"
    worksheet.freeze_panes = cell_address


def set_print_options(worksheet: Worksheet):
    """Set print options for worksheet"""
    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_setup.fitToWidth = 1
    
    # Set print area
    worksheet.print_area = f'A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}'


def apply_alternating_row_colors(worksheet: Worksheet, start_row: int = 2):
    """Apply alternating row colors for better readability"""
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    for row_num in range(start_row, worksheet.max_row + 1):
        if (row_num - start_row) % 2 == 1:
            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.fill = light_fill


def format_number_columns(worksheet: Worksheet, number_columns: List[int], 
                         decimal_places: int = 2):
    """Format specified columns as numbers with given decimal places"""
    for col_idx in number_columns:
        for row in range(2, worksheet.max_row + 1):  # Skip header row
            cell = worksheet.cell(row=row, column=col_idx)
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = f'0.{"0" * decimal_places}'


def add_data_validation(worksheet: Worksheet, column: int, valid_values: List[str],
                       start_row: int = 2):
    """Add data validation (dropdown) to a column"""
    from openpyxl.worksheet.datavalidation import DataValidation
    
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(valid_values)}"',
        allow_blank=True
    )
    
    dv.error = 'Invalid entry'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please select from the list'
    dv.promptTitle = 'Valid Options'
    
    # Apply to column
    from openpyxl.utils import get_column_letter
    col_letter = get_column_letter(column)
    dv.add(f'{col_letter}{start_row}:{col_letter}{worksheet.max_row}')
    
    worksheet.add_data_validation(dv)


def merge_cells_with_same_value(worksheet: Worksheet, column: int, start_row: int = 2):
    """Merge cells with same value in a column"""
    from openpyxl.utils import get_column_letter
    
    col_letter = get_column_letter(column)
    current_value = None
    merge_start = start_row
    
    for row in range(start_row, worksheet.max_row + 2):  # +2 to handle last group
        if row <= worksheet.max_row:
            cell_value = worksheet[f'{col_letter}{row}'].value
        else:
            cell_value = None  # Force merge of last group
        
        if cell_value != current_value:
            if current_value is not None and row - merge_start > 1:
                # Merge cells
                worksheet.merge_cells(f'{col_letter}{merge_start}:{col_letter}{row-1}')
                # Center align merged cell
                worksheet[f'{col_letter}{merge_start}'].alignment = Alignment(
                    horizontal='center', vertical='center'
                )
            
            current_value = cell_value
            merge_start = row


def create_summary_row(worksheet: Worksheet, row: int, label: str, 
                      sum_columns: List[int], bold: bool = True):
    """Create a summary row with totals"""
    from openpyxl.utils import get_column_letter
    
    # Add label
    worksheet.cell(row=row, column=1, value=label)
    
    # Apply formatting
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=row, column=col)
        if bold:
            cell.font = Font(bold=True)
        
        # Add sum formulas for specified columns
        if col in sum_columns:
            col_letter = get_column_letter(col)
            # Assuming data starts from row 2 to row-1
            cell.value = f'=SUM({col_letter}2:{col_letter}{row-1})'
            cell.number_format = '0.00'