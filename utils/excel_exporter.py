"""
Excel Export Utilities
Handles all Excel file generation and formatting
"""
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from collections import defaultdict

from config.settings import COLORS, MAX_SHEET_NAME_LENGTH, MAX_COLUMN_WIDTH, MIN_COLUMN_WIDTH
from models.data_models import TimeEntry, Task


class ExcelExporter:
    """Handles Excel file generation and formatting"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.writer = None
    
    def __enter__(self):
        self.writer = pd.ExcelWriter(self.file_path, engine='openpyxl')
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.writer:
            self.writer.close()
    
    def create_time_summary_sheet(self, time_entries_by_user: Dict[str, List[TimeEntry]]):
        """Create summary sheet for time tracking"""
        summary_data = []
        
        for username, entries in time_entries_by_user.items():
            total_hours = sum(entry.hours for entry in entries)
            user_email = entries[0].email if entries else ''
            
            summary_data.append({
                'User': username,
                'Email': user_email,
                'Total Entries': len(entries),
                'Total Hours': round(total_hours, 2)
            })
        
        # Add total row
        total_entries = sum(len(entries) for entries in time_entries_by_user.values())
        total_hours = sum(
            sum(entry.hours for entry in entries) 
            for entries in time_entries_by_user.values()
        )
        
        summary_data.append({
            'User': 'TOTAL',
            'Email': '',
            'Total Entries': total_entries,
            'Total Hours': round(total_hours, 2)
        })
        
        df = pd.DataFrame(summary_data)
        df.to_excel(self.writer, sheet_name='Summary', index=False)
        self._format_header(self.writer, 'Summary')
        self._auto_adjust_columns('Summary')
    
    def create_time_data_sheet(self, time_entries_by_user: Dict[str, List[TimeEntry]]):
        """Create sheet with all time tracking data"""
        all_data = []
        
        for username, entries in time_entries_by_user.items():
            for entry in entries:
                all_data.append(entry.to_dict())
        
        if all_data:
            df = pd.DataFrame(all_data)
            df = df.sort_values(['日期', '員工'])
            df.to_excel(self.writer, sheet_name='All Data', index=False)
            
            self._format_header(self.writer, 'All Data')
            self._auto_adjust_columns('All Data')
            self._make_urls_clickable('All Data', '連結')
    
    def create_user_time_sheets(self, time_entries_by_user: Dict[str, List[TimeEntry]]):
        """Create individual user time sheets"""
        for username, entries in time_entries_by_user.items():
            safe_name = self._make_safe_sheet_name(username)
            
            if entries:
                user_data = [entry.to_dict() for entry in entries]
                df = pd.DataFrame(user_data)
                df = df.sort_values('日期')
            else:
                df = pd.DataFrame(columns=[
                    '日期', '員工', '專案', '專項', '任務', '描述', '工時', '連結'
                ])
            
            df.to_excel(self.writer, sheet_name=safe_name, index=False)
            self._format_header(self.writer, safe_name)
            self._auto_adjust_columns(safe_name)
            if '連結' in df.columns:
                self._make_urls_clickable(safe_name, '連結')
    
    def create_task_summary_sheet(self, tasks_by_folder: Dict[str, List[Task]]):
        """Create summary sheet for tasks"""
        summary_data = []
        
        for folder_name, tasks in tasks_by_folder.items():
            if not tasks:
                continue
            
            # Count tasks by status
            status_counts = defaultdict(int)
            total_time_spent = 0
            total_time_estimated = 0
            total_points = 0
            
            for task in tasks:
                status_counts[task.status.lower()] += 1
                total_time_spent += task.time_spent
                total_time_estimated += task.time_estimate
                total_points += task.points
            
            # Group by list to count lists
            tasks_by_list = defaultdict(list)
            for task in tasks:
                tasks_by_list[task.list_name].append(task)
            
            summary_data.append({
                'Folder': folder_name,
                'Total Tasks': len(tasks),
                'Open Tasks': (status_counts.get('open', 0) + 
                              status_counts.get('to do', 0) + 
                              status_counts.get('backlog', 0)),
                'In Progress': (status_counts.get('in progress', 0) + 
                               status_counts.get('in development', 0)),
                'Completed': (status_counts.get('complete', 0) + 
                             status_counts.get('closed', 0)),
                'Hours Spent': round(total_time_spent / (1000 * 60 * 60), 2) if total_time_spent else 0,
                'Hours Estimated': round(total_time_estimated / (1000 * 60 * 60), 2) if total_time_estimated else 0,
                'Sprint Points': total_points,
                'Lists Count': len(tasks_by_list)
            })
        
        # Add grand total
        if summary_data:
            grand_total = {
                'Folder': 'GRAND TOTAL',
                'Total Tasks': sum(row['Total Tasks'] for row in summary_data),
                'Open Tasks': sum(row['Open Tasks'] for row in summary_data),
                'In Progress': sum(row['In Progress'] for row in summary_data),
                'Completed': sum(row['Completed'] for row in summary_data),
                'Hours Spent': sum(row['Hours Spent'] for row in summary_data),
                'Hours Estimated': sum(row['Hours Estimated'] for row in summary_data),
                'Sprint Points': sum(row['Sprint Points'] for row in summary_data),
                'Lists Count': sum(row['Lists Count'] for row in summary_data)
            }
            summary_data.append(grand_total)
        
        df = pd.DataFrame(summary_data)
        df.to_excel(self.writer, sheet_name='Summary', index=False)
        self._format_header(self.writer, 'Summary')
        self._auto_adjust_columns('Summary')
    
    def create_folder_task_sheet(self, folder_name: str, tasks: List[Task], 
                                include_custom_fields: bool = True,
                                include_time_tracking: bool = True,
                                include_assignees: bool = True,
                                include_dates: bool = True):
        """Create sheet for folder tasks organized by lists"""
        safe_name = self._make_safe_sheet_name(folder_name)
        
        # Group tasks by list
        tasks_by_list = defaultdict(list)
        for task in tasks:
            tasks_by_list[task.list_name].append(task)
        
        all_task_data = []
        folder_total_hours = 0
        folder_total_points = 0
        
        # Process each list
        for list_name in sorted(tasks_by_list.keys()):
            list_tasks = tasks_by_list[list_name]
            
            # Calculate list totals
            list_total_hours = sum(task.hours_spent for task in list_tasks)
            list_total_points = sum(task.points for task in list_tasks)
            folder_total_hours += list_total_hours
            folder_total_points += list_total_points
            
            # Add list header
            list_header_text = f"=== {list_name} ({len(list_tasks)} tasks"
            if include_time_tracking and (list_total_hours > 0 or list_total_points > 0):
                list_header_text += f" | {list_total_hours}h"
                if list_total_points > 0:
                    list_header_text += f" | {list_total_points}pts"
            list_header_text += ") ==="
            
            # Create list header row
            list_row = self._create_task_row_template(
                include_custom_fields, include_time_tracking, 
                include_assignees, include_dates, list_tasks
            )
            list_row['List'] = list_header_text
            if include_time_tracking:
                list_row['Hours Spent'] = f"Total: {list_total_hours}h"
                list_row['Sprint Points'] = f"Total: {list_total_points}pts" if list_total_points > 0 else ''
            
            all_task_data.append(list_row)
            
            # Add tasks
            for task in list_tasks:
                task_row = self._create_task_row(
                    task, include_custom_fields, include_time_tracking,
                    include_assignees, include_dates
                )
                all_task_data.append(task_row)
            
            # Add separator
            all_task_data.append({col: '' for col in list_row.keys()})
        
        # Remove last separator
        if all_task_data and all(v == '' for v in all_task_data[-1].values()):
            all_task_data.pop()
        
        # Add folder total at beginning
        if include_time_tracking and (folder_total_hours > 0 or folder_total_points > 0):
            folder_total_row = {col: '' for col in all_task_data[0].keys() if all_task_data}
            folder_total_row['List'] = f"FOLDER TOTAL: {folder_total_hours}h"
            if folder_total_points > 0:
                folder_total_row['List'] += f" | {folder_total_points}pts"
            if include_time_tracking:
                folder_total_row['Hours Spent'] = f"{folder_total_hours}h"
                folder_total_row['Sprint Points'] = f"{folder_total_points}pts" if folder_total_points > 0 else ''
            
            all_task_data.insert(0, folder_total_row)
            all_task_data.insert(1, {col: '' for col in folder_total_row.keys()})
        
        df = pd.DataFrame(all_task_data)
        df.to_excel(self.writer, sheet_name=safe_name, index=False)
        
        self._format_task_sheet_with_lists(safe_name)
        self._auto_adjust_columns(safe_name)
        if 'Task URL' in df.columns:
            self._make_urls_clickable(safe_name, 'Task URL')
    
    def _create_task_row_template(self, include_custom_fields: bool, include_time_tracking: bool,
                                 include_assignees: bool, include_dates: bool, tasks: List[Task]) -> Dict:
        """Create template row with all columns"""
        row = {
            'Task ID': '',
            'List': '',
            'Task Name': '',
            'Status': '',
            'Description': ''
        }
        
        if include_assignees:
            row['Assignees'] = ''
        
        if include_dates:
            row['Due Date'] = ''
            row['Start Date'] = ''
        
        if include_time_tracking:
            row['Hours Spent'] = ''
            row['Hours Estimated'] = ''
            row['Sprint Points'] = ''
        
        if include_custom_fields:
            # Get all custom field names
            custom_field_names = set()
            for task in tasks:
                for field_name in task.custom_fields.keys():
                    custom_field_names.add(f"CF: {field_name}")
            
            for field_name in sorted(custom_field_names):
                row[field_name] = ''
        
        row['Task URL'] = ''
        return row
    
    def _create_task_row(self, task: Task, include_custom_fields: bool, include_time_tracking: bool,
                        include_assignees: bool, include_dates: bool) -> Dict:
        """Create row data for a task"""
        row = {
            'Task ID': task.id,
            'List': task.list_name,
            'Task Name': task.name,
            'Status': task.status,
            'Description': task.description
        }
        
        if include_assignees:
            row['Assignees'] = ', '.join(task.assignees)
        
        if include_dates:
            row['Due Date'] = task.due_date or ''
            row['Start Date'] = task.start_date or ''
        
        if include_time_tracking:
            row['Hours Spent'] = task.hours_spent
            row['Hours Estimated'] = task.hours_estimated
            row['Sprint Points'] = task.points
        
        if include_custom_fields:
            for field_name, field_value in task.custom_fields.items():
                row[f"CF: {field_name}"] = field_value
        
        row['Task URL'] = task.url
        return row
    
    def _make_safe_sheet_name(self, name: str) -> str:
        """Create Excel-safe sheet name"""
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        safe_name = str(name)
        for char in invalid_chars:
            safe_name = safe_name.replace(char, '_')
        
        if len(safe_name) > MAX_SHEET_NAME_LENGTH:
            safe_name = safe_name[:MAX_SHEET_NAME_LENGTH-3] + "..."
        
        return safe_name
    
    def _format_header(self, writer, sheet_name: str):
        """Format header row"""
        try:
            worksheet = writer.sheets[sheet_name]
            header_fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type="solid")
            header_font = Font(color=COLORS['header_text'], bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
        except Exception as e:
            print(f"Warning: Could not format header for {sheet_name}: {e}")
    
    def _format_task_sheet_with_lists(self, sheet_name: str):
        """Format task sheet with list headers highlighted"""
        try:
            worksheet = self.writer.sheets[sheet_name]
            
            # Header formatting
            header_fill = PatternFill(start_color=COLORS['task_header'], end_color=COLORS['task_header'], fill_type="solid")
            header_font = Font(color=COLORS['task_header_text'], bold=True)
            
            # List header formatting
            list_header_fill = PatternFill(start_color=COLORS['list_header'], end_color=COLORS['list_header'], fill_type="solid")
            list_header_font = Font(color=COLORS['list_header_text'], bold=True)
            
            # Format main header
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Format list headers (rows that contain "=== List Name ===")
            for row in worksheet.iter_rows(min_row=2):
                if row[1].value and str(row[1].value).startswith("==="):
                    for cell in row:
                        cell.fill = list_header_fill
                        cell.font = list_header_font
                        cell.alignment = Alignment(horizontal="left")
        except Exception as e:
            print(f"Warning: Could not format task sheet {sheet_name}: {e}")
    
    def _auto_adjust_columns(self, sheet_name: str):
        """Auto-adjust column widths"""
        try:
            worksheet = self.writer.sheets[sheet_name]
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        cell_length = self._calculate_display_width(str(cell.value))
                        max_length = max(max_length, cell_length)
                
                adjusted_width = min(max(max_length + 2, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            print(f"Warning: Could not adjust columns for {sheet_name}: {e}")
    
    def _calculate_display_width(self, text: str) -> int:
        """Calculate display width (Chinese chars = 2, others = 1)"""
        width = 0
        for char in text:
            if ord(char) > 127:  # Non-ASCII (including Chinese)
                width += 2
            else:
                width += 1
        return width
    
    def _make_urls_clickable(self, sheet_name: str, url_column: str):
        """Make URLs clickable in Excel"""
        try:
            worksheet = self.writer.sheets[sheet_name]
            
            # Find URL column
            url_col_index = None
            for col in worksheet[1]:  # Header row
                if col.value == url_column:
                    url_col_index = col.column
                    break
            
            if url_col_index is None:
                return
            
            # Make URLs clickable
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=url_col_index)
                if cell.value and str(cell.value).startswith('http'):
                    cell.hyperlink = cell.value
                    cell.font = Font(color=COLORS['link'], underline="single")
                    cell.value = "Link"
        except Exception as e:
            print(f"Warning: Could not make URLs clickable for {sheet_name}: {e}")